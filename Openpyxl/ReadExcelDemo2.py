import openpyxl

wk=openpyxl.load_workbook("F:\Excel Automation\ReadExcel.xlsx")
sh=wk["Members"]

#Case1
print(sh['A1'].value)
print(sh['B3'].value)

#Case2
c1=sh.cell(1,2)
print(c1.value)

#Case3
c1=sh.cell(column=1,row=2)
print(c1.value)

print(c1.row)
print(c1.column)