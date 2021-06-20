import openpyxl

wk=openpyxl.load_workbook("F:\Excel Automation\ReadExcel.xlsx")
sh=wk["Members"]

# To print all the rows and columns
row=sh.max_row
column=sh.max_column

print("Total rows are " + str(row))
print("Total columns are "+ str(column))

# first approch
#for i in range(1, row+1):
#    for j in range(1,column+1):
#       c=sh.cell(i,j)
#        print(c.value)

# second approch
for r in sh['A1':'C3']:
    for c in r:
        print(c.value)